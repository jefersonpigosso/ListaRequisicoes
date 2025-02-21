import streamlit as st
from streamlit import session_state as sts
import warnings
import pandas  as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side

################################### Configurações Iniciais ###################################

warnings.filterwarnings("ignore", message="Workbook contains no default style, apply openpyxl's default")

st.set_page_config(layout='wide')

################################### Funções ###################################

def requisicoes(requisicoes, enderecamento):

    # cria as tabelas com os arquivos
    tabRequisicoes      = pd.read_excel(
                            requisicoes,
                            engine='openpyxl', dtype=object, usecols=['Produto', 'Tamanho Pack', 'Volumes', 'Status'])
    tabEnderecamento    = pd.read_excel(
                            enderecamento,
                            engine='openpyxl', dtype=object, usecols=['Produto', 'Peças', 'Endereço', 'UZ', 'Fase'])

    # renomeia as colunas das tabelas
    tabRequisicoes      = tabRequisicoes.rename(
                            columns={
                                'Produto'       :   'cod_produto',
                                'Tamanho Pack'  :   'grade',
                                'Volumes'       :   'volumes',
                                'Status'        :   'status'
                            })
    tabEnderecamento    = tabEnderecamento.rename(
                            columns={
                                'Produto'   :   'cod_produto',
                                'Peças'     :   'grade',
                                'Endereço'  :   'endereco_palete',
                                'UZ'        :   'endereco_uz',
                                'Fase'      :   'fase'
                            })                    

    # filtra as tabelas
    tabRequisicoes      = tabRequisicoes[tabRequisicoes['status'] != 'FINALIZADO']
    tabEnderecamento    = tabEnderecamento[tabEnderecamento['fase'] == 'Armazenagem']

    # remove espaco em branco da coluna
    tabRequisicoes['cod_produto']   = tabRequisicoes['cod_produto'].str.rstrip()
    tabEnderecamento['cod_produto'] = tabEnderecamento['cod_produto'].str.rstrip()

    # agrupa as tabelas
    tabRequisicoes      = tabRequisicoes.groupby(
                            ['cod_produto', 'grade'],
                            dropna=False,
                            as_index=False)\
                            .agg({'volumes': 'sum'})
    tabEnderecamento    = tabEnderecamento.groupby(
                            ['cod_produto', 'grade', 'endereco_palete', 'endereco_uz', 'fase'],
                            dropna=False,
                            as_index=False)\
                            .agg({'fase': 'count'})\
                            .rename(columns={'fase': 'volumes'})

    # ordena as tabelas
    tabRequisicoes      =      tabRequisicoes.sort_values(
                                    by=['cod_produto', 'grade', 'volumes'],
                                    ascending=[True, True, True])
    tabEnderecamento    =      tabEnderecamento.sort_values(
                                    by=['cod_produto', 'grade', 'volumes', 'endereco_palete', 'endereco_uz'],
                                    ascending=[True, True, True, True, True])

    # cria uma unica coluna de endereco
    tabEnderecamento['endereco'] = tabEnderecamento['endereco_palete'].astype(str) + '-' + tabEnderecamento['endereco_uz'].astype(str) + '-' + tabEnderecamento['volumes'].astype(str)

    ########################################################################################

    # cria tabela com as posicoes unicas
    tabUnico    =   tabRequisicoes.merge(
                        tabEnderecamento,
                        left_on=['cod_produto', 'grade'],
                        right_on=['cod_produto', 'grade'],
                        how='left',
                        validate=('1:m'))\
                        .rename(columns={
                            'volumes_x'         :   'volumes',
                            'volumes_y'         :   'vol_acumulado'})

    # filtra a tabela pelos volumes nos endereços maiores que a requisicao
    tabUnico    =   tabUnico[tabUnico['vol_acumulado'] >= tabUnico['volumes']]

    # remove as duplicatas da tabela e mantem o primeiro valor
    tabUnico    =   tabUnico\
                        .drop_duplicates(subset=['cod_produto', 'grade'])

    ########################################################################################

    # cria tabela com as posicoes unicas
    tabMultiplo =   tabRequisicoes.merge(
                        tabUnico,
                        left_on=['cod_produto', 'grade'],
                        right_on=['cod_produto', 'grade'],
                        how='left',
                        validate=('1:m'),
                        indicator=True)

    # filtra a tabela para mostrar somente os que precisam de mais de uma posicao
    tabMultiplo =   tabMultiplo[tabMultiplo['_merge'] == 'left_only']

    # configura a tabela
    tabMultiplo =   tabMultiplo\
                        .rename (columns={'volumes_x': 'volumes'})\
                        .drop   (columns=['volumes_y', '_merge'])\
                        .sort_values(   by          =['cod_produto', 'grade', 'volumes'],
                                        ascending   =[True, False, False])

    # cria coluna para volume acumulado
    tabMultiplo['vol_acumulado'] = 0

    # loop pela tabela multiplo
    for produto in tabMultiplo['cod_produto'].index:
        
        # cria variaveis para buscar na tabela enderecamento
        cod_produto_multiplo    =   tabMultiplo.loc[produto, 'cod_produto']
        grade_multiplo          =   tabMultiplo.loc[produto, 'grade']
        volumes_multiplo        =   tabMultiplo.loc[produto, 'volumes']

        # cria tabela de enderecamento filtrado pelo codigo do produto e grade atual
        tabEnderecamentoFiltrado    =   tabEnderecamento[
                                            (tabEnderecamento['cod_produto']    == cod_produto_multiplo) &
                                            (tabEnderecamento['grade']          == grade_multiplo)]

        # ordena a tabela para ter o maior volume primeiro
        tabEnderecamentoFiltrado    =   tabEnderecamentoFiltrado\
                                            .sort_values(   by          =['cod_produto', 'grade', 'volumes', 'endereco'],
                                                            ascending   =[True, True, False, True])

        # verifica se a tabela enderecamento filtrado nao está vazia
        if not tabEnderecamentoFiltrado.empty:
            
            # cria variaveis para acumular os valores da tabela enderecamento
            endereco_acumulado                  =   ''
            volumes_enderecamento_acumulado =   0

            # validação do codigo para pular os erros
            try:

                # loop pela tabela de enderecamento filtrado (buscando os volumes suficientes para a requisicao)
                for endereco in tabEnderecamentoFiltrado.index:
                    
                    # cria variaveis para buscar na tabela enderecamento
                    endereco_enderecamento  =   str(tabEnderecamentoFiltrado.loc[endereco, 'endereco'])
                    volumes_enderecamento   =   tabEnderecamentoFiltrado.loc[endereco, 'volumes']

                    # acumula os valores das variaveis numa unica variavel
                    endereco_acumulado              +=  f'{endereco_enderecamento}\n'
                    volumes_enderecamento_acumulado +=  volumes_enderecamento

                    # adiciona a o endereco para a requisicao
                    tabMultiplo.loc[produto, 'endereco'] = endereco_acumulado

                    # verifica se o volume acumulado é maior que o volume da requisicao
                    if volumes_enderecamento_acumulado >= volumes_multiplo:
                        
                        # insere o volume acumulado na coluna (usar como filtro posteriormente)
                        tabMultiplo.loc[produto, 'vol_acumulado'] = volumes_enderecamento_acumulado

                        # encerra o loop atual
                        break
            
            # caso não haja volumes enderecados suficientes para atender a requisição, irá para a proxima e deixará a atual em branco
            except:
                pass

    ########################################################################################

    # cria a tabela com o arquivo
    requisicoes      = pd.read_excel(
                            requisicoes,
                            engine='openpyxl', dtype=object, usecols=['Autorizacao', 'Status', 'Produto', 'Griffe', 'Volumes', 'Tamanho Pack', 'Peças', 'Solicitação', 'Linha'])

    # renomeia as colunas da tabela
    requisicoes      = requisicoes.rename(
                            columns={
                                'Autorizacao'   :   'requisicao',
                                'Status'        :   'status',
                                'Produto'       :   'cod_produto',
                                'Griffe'        :   'grife',                            
                                'Volumes'       :   'volumes',
                                'Tamanho Pack'  :   'grade',
                                'Peças'         :   'pecas',
                                'Solicitação'   :   'data_criacao',
                                'Linha'         :   'linha'
                            })

    # filtra a tabela
    requisicoes = requisicoes[requisicoes['status'] != 'FINALIZADO']

    # remove as horas da coluna data
    requisicoes['data_criacao'] = requisicoes['data_criacao'].astype(str).str[:10]

    # transforma as duas tabelas em uma só
    tabRequisicoes = pd.concat([tabUnico, tabMultiplo], ignore_index=True)

    # altera o tipo das colunas
    requisicoes[['cod_produto', 'grade']]       = requisicoes[['cod_produto', 'grade']].astype(str)
    tabRequisicoes[['cod_produto', 'grade']]    = tabRequisicoes[['cod_produto', 'grade']].astype(str)

    # remove espaco em branco da coluna
    requisicoes['cod_produto']  = requisicoes['cod_produto'].str.rstrip()
    tabRequisicoes['endereco']  = tabRequisicoes['endereco'].str.rstrip()
    requisicoes['grife']        = requisicoes['grife'].str.rstrip()
    requisicoes['linha']        = requisicoes['linha'].str.rstrip()

    # mescla as duas tabelas em uma
    requisicoes = requisicoes\
                    .merge(
                        tabRequisicoes,
                        left_on=['cod_produto', 'grade'],
                        right_on=['cod_produto', 'grade'],
                        how='left',
                        validate=('m:1'))\
                    .drop(
                        columns=['volumes_y'])\
                    .rename(
                        columns={
                            'volumes_x' : 'volumes'})

    # transforma as colunas
    requisicoes['status']   =   requisicoes['status'].str.title()
    requisicoes['grife']    =   requisicoes['grife'].str.title()
    requisicoes['linha']    =   requisicoes['linha'].str.title()

    # remove o texto nan das colunas
    requisicoes['endereco']          = requisicoes['endereco'].str.replace('nan', '').fillna('') # \A busca somente no começo do texto, precisa explicitar o regex para true

    # insere linhas aonde houver mais de um endereco (separado por \n (nova linha na celula))
    def convert_to_list(row):
        arr = row.split('\n')
        l = [x for x in arr]
        return l
    requisicoes['endereco'] = requisicoes['endereco'].astype(str).apply(convert_to_list)
    requisicoes = requisicoes.explode('endereco')

    # transforma a coluna endereco em 3 colunas
    requisicoes[['endereco_palete', 'endereco_uz', 'volumes_enderecamento']] = requisicoes['endereco'].str.split('-', n=2, expand=True)

    # cria coluna para filtrar os enderecados e nao enderecados
    requisicoes['endereco'] = requisicoes['endereco_palete'].astype(str).apply(len)

    # altera o tipo das colunas
    requisicoes['volumes_enderecamento'] = requisicoes['volumes_enderecamento'].astype(float)

    # cria colunas para inserção manual
    requisicoes['operador']         = ''
    requisicoes['tipo_requisicao']  = ''

    # Remove insuficientes
    requisicoes = requisicoes.query('vol_acumulado != 0 & endereco_uz.notna()').drop(columns=['vol_acumulado', 'endereco'])

    return requisicoes

def lista_requisicoes(df):

    # Reseta o indice do df para nao bugar no loop
    requisicoes = df.reset_index(drop=True)

    # loop pela planilha para substituir outras posicoes alem do PP
    # se tiver outra posição alem do PP para MANTER, inserir na tupla
    # for endereco in requisicoes['requisicao'].index:
    #     if not str(requisicoes.loc[endereco, 'endereco_palete']).startswith(tuple(['PP'])):
    #         requisicoes.loc[endereco, 'endereco_palete'] = ''

    # obtem os dados do local da posição palete
    requisicoes['rua']          = requisicoes['endereco_palete'].str.slice(2,4)
    requisicoes['local_rua']    = requisicoes['endereco_palete'].str.slice(4,6)
    requisicoes['andar']        = requisicoes['endereco_palete'].str.slice(7,10)

    # converte os valores em branco das posicoes para um numero fixo (usado para nao dar erro ao importar para o sharepoint)
    requisicoes['rua']          = requisicoes['rua']        .astype(str).replace(['nan',''], '9999')
    requisicoes['local_rua']    = requisicoes['local_rua']  .astype(str).replace(['nan',''], '9999')
    requisicoes['andar']        = requisicoes['andar']      .astype(str).replace(['nan',''], '9999')

    # cria coluna id (para nao duplicar no sharepoint)
    requisicoes['key_id'] = \
        requisicoes['requisicao'].astype(str) + requisicoes['cod_produto'].astype(str) + requisicoes['grade'].astype(str) +\
        requisicoes['data_criacao'].astype(str) + requisicoes['endereco_palete'].astype(str) + requisicoes['endereco_uz'].astype(str)

    ############################################################################################################

    # Seleciona somente as colunas necessarias do dataframe requisicoes
    tbRequisicoes = requisicoes[['requisicao', 'cod_produto', 'grade', 'endereco_palete', 'endereco_uz', 'volumes', 'volumes_enderecamento', 'data_criacao', 'grife']]

    # Preenche os dados vazios das colunas evitando mensagem de copiar numa parte do dataframe usando loc
    tbRequisicoes.loc[:, 'endereco_palete'] = tbRequisicoes['endereco_palete'].fillna('-')
    tbRequisicoes.loc[:, 'endereco_uz'] = tbRequisicoes['endereco_uz'].fillna('-')

    tbRequisicoes = tbRequisicoes.sort_values(
            by=['volumes', 'endereco_palete', 'endereco_uz'],
            ascending=[False, True, True]
        ).rename(
            columns={
                'data_criacao' : 'Data',
                'requisicao' : 'Requisição',
                'cod_produto' : 'Produto',
                'grade' : 'Grade',
                'endereco_palete' : 'Posição',
                'endereco_uz' : 'UZ',
                'volumes_enderecamento' : 'Vol',
                'volumes' : 'Total',
                'grife': 'Grife'
                
            }
        )

    tbRequisicoes = tbRequisicoes.reindex(columns=['Data', 'Requisição', 'Produto', 'Grade', 'Posição', 'UZ', 'Vol', 'Total', 'Grife'])

    tbRequisicoes['Data'] = tbRequisicoes['Data'].astype(str)
    tbRequisicoes['Requisição'] = tbRequisicoes['Requisição'].astype(str)
    tbRequisicoes['Produto'] = tbRequisicoes['Produto'].astype(str)
    tbRequisicoes['Grade'] = tbRequisicoes['Grade'].astype(str)
    tbRequisicoes['Vol'] = tbRequisicoes['Vol'].astype(int)
    tbRequisicoes['Vol'] = tbRequisicoes['Vol'].astype(str)
    tbRequisicoes['Vol Ende.'] = tbRequisicoes['Total'].astype(str)
    tbRequisicoes['Grife'] = tbRequisicoes['Grife'].astype(str)

    tbRequisicoes['Data'] = tbRequisicoes['Data'].str.slice(0, 5)

    tbRequisicoesGroup = tbRequisicoes.groupby(
        ['Data', 'Requisição', 'Produto', 'Grade', 'Total', 'Grife'], dropna=False, as_index=False
    ).agg(
        {
            'Posição': '\n'.join,
            'UZ': '\n'.join,
            'Vol': '\n'.join,
        }
    )

    tbRequisicoesGroup = tbRequisicoesGroup[['Data', 'Requisição', 'Produto', 'Grade', 'Posição', 'UZ', 'Vol', 'Total', 'Grife']]

    # Cria uma planilha virtual para baixar em formato excel
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine='xlsxwriter') as writer:
        tbRequisicoesGroup.to_excel(writer, sheet_name='Requisições', index=False)

    buffer.seek(0)
    wb = load_workbook(buffer)
    ws = wb.active

    # Formata o excel
    for row in ws.iter_rows(min_row=1, max_col=ws.max_column, max_row=ws.max_row):
        for cell in row:
            if cell.column in (5, 6, 7):
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal='center', vertical='center')
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                top=Side(style='thin'), bottom=Side(style='thin'))
            cell.border = thin_border

    # Ajustar larguras das colunas
    column_widths = {'A': 10, 'B': 12, 'C': 15, 'D': 9, 'E': 17, 'F': 17, 'G': 9, 'H': 9, 'I': 17}
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Criar um novo buffer para salvar o arquivo final
    final_buffer = BytesIO()
    wb.save(final_buffer)
    final_buffer.seek(0)

    return final_buffer

################################### Layout ###################################

col1, col2 = st.columns(2, gap='large')
with col1:
    st.file_uploader('Requisições', key='df_requisicoes')
with col2:
    st.file_uploader('Endereçamento', key='df_enderecamento')

# Botão para gerar a lista de requisições
st.button(
    label='Gerar lista de requisições',
    key='processar',
    use_container_width=True,
    disabled= not(sts.df_requisicoes is not None and sts.df_enderecamento is not None)
)

################################### Interações ###################################

if sts.processar:

    # Executa o codigo para gerar lista de requisicoes
    df = requisicoes(sts.df_requisicoes, sts.df_enderecamento)
    lista = lista_requisicoes(df)

    # Botão de download da lista
    st.download_button(label="Baixar lista", data=lista, file_name='Lista Requisições.xlsx', mime='application/vnd.ms-excel')

